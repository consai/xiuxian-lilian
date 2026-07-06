#!/usr/bin/env python3
"""将 Excel 源表中的大境界英文 id 替换为拼音 id（保留大道领域 foundation 等非境界用法）。"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: py -3 -m pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1] / "xiuxian配置表"

REALM_MAP: dict[str, str] = {
    "qi": "lianqi",
    "foundation": "zhuji",
    "core": "jindan",
    "nascent": "yuanying",
    "transform": "huashen",
    "void": "lianxu",
    "merge": "heti",
    "great": "dacheng",
    "tribulation": "dujie",
}

TRANSITION_MAP: dict[str, str] = {
    "qi_to_foundation": "lianqi_to_zhuji",
    "foundation_to_core": "zhuji_to_jindan",
    "core_to_nascent": "jindan_to_yuanying",
}

REALM_FIELD_HEADERS = {
    "realm",
    "req_realm",
    "major_realm",
    "from_major",
    "to_major",
    "anchor_realm",
}

REALM_ID_SHEETS = {
    "z_dao_tree_realms",
    "z_jingjie_balance_major_realms",
}

SKIP_ID_SHEETS = {
    "z_dao_tree_domains",
    "z_dao_tree_domaingroups",
}

DOMAIN_SKILL_ID = re.compile(r"^[a-z_]+\.[a-z0-9_]+$", re.I)
JSON_REALM_FIELD = re.compile(
    r'"(realm|major_realm|req_realm|from_major|to_major|anchor_realm)"\s*:\s*"(qi|foundation|core|nascent|transform|void|merge|great|tribulation)"'
)
COMPOUND_PREFIXES = sorted(REALM_MAP.keys(), key=len, reverse=True)


def map_realm_token(token: str) -> str:
    return REALM_MAP.get(token, token)


def map_compound_key(key: str) -> str:
    if key in TRANSITION_MAP:
        return TRANSITION_MAP[key]
    if key in REALM_MAP:
        return REALM_MAP[key]
    for old in COMPOUND_PREFIXES:
        prefix = f"{old}_"
        if key == old:
            return REALM_MAP[old]
        if key.startswith(prefix):
            return REALM_MAP[old] + key[len(old) :]
    if key.startswith("major_realm_multipliers:"):
        suffix = key.split(":", 1)[1]
        if suffix in REALM_MAP:
            return f"major_realm_multipliers:{REALM_MAP[suffix]}"
    return key


def replace_json_realms(text: str) -> tuple[str, int]:
    count = 0

    def _sub(match: re.Match[str]) -> str:
        nonlocal count
        field = match.group(1)
        old = match.group(2)
        new = REALM_MAP.get(old, old)
        if new != old:
            count += 1
        return f'"{field}": "{new}"'

    out = JSON_REALM_FIELD.sub(_sub, text)

    if '"tier_major_realm"' in out:
        try:
            # 仅替换 tier_major_realm 对象内的境界值
            obj_match = re.search(r'"tier_major_realm"\s*:\s*(\{[^}]+\})', out)
            if obj_match:
                raw = obj_match.group(1)
                data = json.loads(raw)
                changed = False
                for k, v in list(data.items()):
                    if isinstance(v, str) and v in REALM_MAP:
                        data[k] = REALM_MAP[v]
                        changed = True
                if changed:
                    new_raw = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
                    out = out.replace(raw, new_raw, 1)
                    count += 1
        except json.JSONDecodeError:
            pass

    return out, count


def should_replace_id_cell(sheet_name: str, value: str) -> bool:
    sheet = sheet_name.lower()
    if sheet in SKIP_ID_SHEETS:
        return False
    if DOMAIN_SKILL_ID.match(value):
        return False
    if sheet == "z_dao_tree_skills":
        return False
    if sheet in REALM_ID_SHEETS:
        return value in REALM_MAP
    return False


def process_cell(sheet_name: str, header: str | None, value: str) -> tuple[str, int]:
    if header in REALM_FIELD_HEADERS and value in REALM_MAP:
        return REALM_MAP[value], 1

    if header in {"key", "id"} or header is None:
        mapped = map_compound_key(value)
        if mapped != value:
            return mapped, 1

    if header == "id" and should_replace_id_cell(sheet_name, value) and value in REALM_MAP:
        return REALM_MAP[value], 1

    # 宽表列头（row1/2 的 qi、foundation 等）
    if header in REALM_MAP and value == header:
        return REALM_MAP[header], 1

    if "{" in value or "[" in value:
        new_val, n = replace_json_realms(value)
        if n:
            return new_val, n
        # 地点 tags 等纯数组：["resource","foundation"]
        if value.startswith("[") and value.endswith("]"):
            try:
                arr = json.loads(value)
                if isinstance(arr, list):
                    changed = False
                    new_arr = []
                    for item in arr:
                        if isinstance(item, str) and item in REALM_MAP:
                            new_arr.append(REALM_MAP[item])
                            changed = True
                        else:
                            new_arr.append(item)
                    if changed:
                        return json.dumps(new_arr, ensure_ascii=False), 1
            except json.JSONDecodeError:
                pass

    # major_realm_multipliers:qi 等整格字符串
    if value.startswith("major_realm_multipliers:"):
        mapped = map_compound_key(value)
        if mapped != value:
            return mapped, 1

    if value in TRANSITION_MAP:
        return TRANSITION_MAP[value], 1

    return value, 0


def header_name(cell_value) -> str | None:
    if cell_value is None:
        return None
    return str(cell_value).strip().lower()


def process_workbook(path: Path) -> int:
    wb = openpyxl.load_workbook(path)
    total = 0
    for ws in wb.worksheets:
        headers: list[str | None] = []
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            if row_idx <= 3:
                # 列头行：把 qi/foundation 等宽表字段名也替换
                for col_idx, cell in enumerate(row):
                    if cell.value is None or not isinstance(cell.value, str):
                        continue
                    hdr = header_name(cell.value)
                    if hdr in REALM_MAP:
                        cell.value = REALM_MAP[hdr]
                        total += 1
                    elif hdr in TRANSITION_MAP:
                        cell.value = TRANSITION_MAP[hdr]
                        total += 1
                    else:
                        mapped = map_compound_key(cell.value)
                        if mapped != cell.value:
                            cell.value = mapped
                            total += 1
                if row_idx == 2:
                    headers = [header_name(c.value) for c in row]
                continue

            for col_idx, cell in enumerate(row):
                if cell.value is None or not isinstance(cell.value, str):
                    continue
                hdr = headers[col_idx] if col_idx < len(headers) else None
                # category 列的 tribulation 是效果分类，不替换
                if hdr == "category" and cell.value == "tribulation":
                    continue
                new_val, n = process_cell(ws.title, hdr, cell.value)
                if n:
                    cell.value = new_val
                    total += n

    if total:
        wb.save(path)
    return total


def fix_embedded_json_workbook(path: Path) -> int:
    """补漏：tier_major_realm / major_realm_starts 等 JSON 单元格。"""
    wb = openpyxl.load_workbook(path)
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                text = cell.value.strip()
                if not text.startswith("{") and not text.startswith("["):
                    continue
                try:
                    data = json.loads(text)
                except json.JSONDecodeError:
                    continue
                changed = False

                def map_obj(obj):
                    nonlocal changed
                    if isinstance(obj, dict):
                        out = {}
                        for k, v in obj.items():
                            if k == "id" and isinstance(v, str) and v in REALM_MAP:
                                out[k] = REALM_MAP[v]
                                changed = True
                            elif k in REALM_MAP and isinstance(v, str) and v in REALM_MAP:
                                out[k] = REALM_MAP[v]
                                changed = True
                            else:
                                out[k] = map_obj(v)
                        return out
                    if isinstance(obj, list):
                        return [map_obj(x) for x in obj]
                    if isinstance(obj, str) and obj in REALM_MAP:
                        changed = True
                        return REALM_MAP[obj]
                    return obj

                new_data = map_obj(data)
                if changed:
                    cell.value = json.dumps(new_data, ensure_ascii=False, separators=(",", ":"))
                    total += 1
    if total:
        wb.save(path)
    return total


def main() -> int:
    files = sorted(ROOT.rglob("*.xlsx"))
    grand = 0
    for path in files:
        n = process_workbook(path)
        n += fix_embedded_json_workbook(path)
        if n:
            print(f"updated {path.relative_to(ROOT)} ({n} cells)")
            grand += n
    print(f"done: {grand} replacements in {len(files)} workbooks")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
