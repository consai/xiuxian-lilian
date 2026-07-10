#!/usr/bin/env python3
"""将 Excel 配置表中 json 类型列转为 string[] / string[][] / int[] 等格式。

转换规则：
- JSON 字符串 "xxx"  → 去引号，类型 → string
- JSON 字符串数组     → 用 : 分隔，类型 → string[]
- JSON 数字数组       → 用 : 分隔，类型 → int[]
- JSON 2D 数组        → 用 | : 分隔，类型 → int[][] 或 string[][]
- 复杂对象/嵌套结构   → 保留 JSON 原文，类型 → string（不再自动 json.loads）
- 混合列（标量+数组） → 统一为 string，数组内容用 : 连接为纯文本
"""

import json
import shutil
import sys
from pathlib import Path

import openpyxl

INDIR = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\godot\xiuxian\excel_config\indir")
BACKUP_DIR = INDIR / "_backup_before_json_convert"


def safe_json_parse(value) -> object | None:
    """Attempt to parse as JSON. Returns parsed value or None."""
    if value is None:
        return None
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    if not (text.startswith(("{", "[")) or text.startswith('"')):
        return None
    try:
        return json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return None


def classify(parsed) -> str:
    """Classify a parsed JSON value into a type name."""
    if parsed is None:
        return "null"
    if isinstance(parsed, bool):
        return "bool"
    if isinstance(parsed, int):
        return "int"
    if isinstance(parsed, float):
        return "float"
    if isinstance(parsed, str):
        return "string"
    if isinstance(parsed, list):
        if len(parsed) == 0:
            return "empty_list"
        first = parsed[0]
        if isinstance(first, list):
            return "list2d"
        if isinstance(first, dict):
            return "complex"
        # Check if all same scalar type
        types = {classify(item) for item in parsed}
        if types <= {"string"}:
            return "string_list"
        if types <= {"int"}:
            return "int_list"
        if types <= {"float"}:
            return "float_list"
        if types <= {"int", "float"}:
            return "number_list"
        return "mixed_list"
    if isinstance(parsed, dict):
        return "complex"
    return "unknown"


def convert_cell(parsed):
    """Convert a parsed JSON value to (new_type, new_cell_value).
    Returns (None, None) if cell should be left unchanged."""
    kind = classify(parsed)

    if kind == "null":
        return None, None

    if kind == "string":
        # "hello" → hello
        return "string", parsed

    if kind in ("int", "float", "bool"):
        return kind, parsed

    if kind == "empty_list":
        return "string[]", None  # empty → clear cell

    if kind == "string_list":
        return "string[]", ":".join(str(x) for x in parsed)

    if kind == "int_list":
        return "int[]", ":".join(str(x) for x in parsed)

    if kind in ("float_list", "number_list"):
        return "float[]", ":".join(str(x) for x in parsed)

    if kind == "list2d":
        # Determine inner type
        flat = [item for sub in parsed for item in sub if sub]
        if not flat:
            return "string[][]", json.dumps(parsed, ensure_ascii=False)
        inner_kinds = {classify(item) for item in flat}
        if inner_kinds <= {"int"}:
            inner = "int"
        elif inner_kinds <= {"int", "float"}:
            inner = "float"
        else:
            inner = "string"
        fmt = "|".join(
            ":".join(str(c) for c in row) for row in parsed if row
        )
        return f"{inner}[][]", fmt

    # complex / mixed / unknown → keep as JSON string, type → string
    return "string", json.dumps(parsed, ensure_ascii=False)


def process_workbook(xlsx_path: Path) -> int:
    """
    Process one workbook. Load without data_only so we overwrite
    actual cell values. Returns number of columns modified.
    """
    modified = 0
    wb = openpyxl.load_workbook(xlsx_path)

    for ws in wb.worksheets:
        if ws.max_row < 5:
            continue

        type_row = 3
        field_row = 2
        data_start = 5

        for col_idx in range(1, ws.max_column + 1):
            type_val = ws.cell(row=type_row, column=col_idx).value
            if not type_val or str(type_val).strip().lower() != "json":
                continue

            field_name = ws.cell(row=field_row, column=col_idx).value or f"col_{col_idx}"

            # Collect per-cell conversions
            cell_targets: dict[int, tuple[str, object]] = {}  # row → (type, value)
            type_counts: dict[str, int] = {}
            skip_rows: set[int] = set()

            for row_idx in range(data_start, ws.max_row + 1):
                raw = ws.cell(row=row_idx, column=col_idx).value
                parsed = safe_json_parse(raw)
                if parsed is None:
                    if raw is not None and str(raw).strip():
                        type_counts["string"] = type_counts.get("string", 0) + 1
                    continue
                new_type, new_val = convert_cell(parsed)
                if new_type is None:
                    skip_rows.add(row_idx)
                    continue
                cell_targets[row_idx] = (new_type, new_val)
                type_counts[new_type] = type_counts.get(new_type, 0) + 1

            if not type_counts:
                # Column has no parseable JSON data → change type to string
                ws.cell(row=type_row, column=col_idx).value = "string"
                print(f"  {ws.title}.{field_name} (col {col_idx}): json → string  (no JSON data)")
                modified += 1
                continue

            # Handle mixed columns: if there are multiple distinct types,
            # use string for everything to avoid data loss
            distinct_types = set(type_counts.keys())
            if len(distinct_types) > 1:
                # Mixed column → unify as string
                final_type = "string"
                for row_idx, (_, new_val) in cell_targets.items():
                    if new_val is None:
                        cell_targets[row_idx] = ("string", None)
                    elif isinstance(new_val, str):
                        cell_targets[row_idx] = ("string", new_val)
                    else:
                        cell_targets[row_idx] = ("string", str(new_val))
                print(f"  {ws.title}.{field_name} (col {col_idx}): json → string  "
                      f"(mixed types: {type_counts})")
            else:
                final_type = distinct_types.pop()
                print(f"  {ws.title}.{field_name} (col {col_idx}): json → {final_type}")

            # Write type row
            ws.cell(row=type_row, column=col_idx).value = final_type

            # Write data cells
            for row_idx, (_, new_val) in cell_targets.items():
                ws.cell(row=row_idx, column=col_idx).value = new_val

            modified += 1

    if modified > 0:
        wb.save(xlsx_path)
    wb.close()
    return modified


def main():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    xlsx_files = sorted(
        p for p in INDIR.glob("*.xlsx") if not p.name.startswith("~$")
    )

    if not xlsx_files:
        print("No xlsx files found in indir.")
        return

    total = 0
    for xlsx_path in xlsx_files:
        backup = BACKUP_DIR / xlsx_path.name
        if not backup.exists():
            shutil.copy2(xlsx_path, backup)

        print(f"\n📄 {xlsx_path.name}")
        n = process_workbook(xlsx_path)
        total += n
        print(f"  → {n} column(s) converted" if n else "  → no json columns")

    print(f"\n{'='*60}")
    print(f"Done. {total} column(s) total converted across {len(xlsx_files)} files.")
    print(f"Backups: {BACKUP_DIR}")


if __name__ == "__main__":
    main()
