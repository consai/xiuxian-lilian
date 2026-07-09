#!/usr/bin/env python3
"""按 z_{xxx} 分组拆分 indir 中的 Excel：同名 xxx 的 sheet 归入同一工作簿。"""
from __future__ import annotations

import argparse
import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def group_key(sheet_name: str, all_z_sheets: set[str]) -> str | None:
    """从 z_{xxx} / z_{xxx}_{子表} 命名中解析分组键 xxx。"""
    if not sheet_name.startswith("z_"):
        return None
    rest = sheet_name[2:]
    parts = rest.split("_")
    best = rest
    # 存在根 sheet z_{prefix} 时，子表 z_{prefix}_* 与根表同组
    for i in range(1, len(parts)):
        prefix = "_".join(parts[:i])
        if f"z_{prefix}" in all_z_sheets:
            best = prefix
    return best


def copy_worksheet(source: Worksheet, target_wb: Workbook, title: str) -> Worksheet:
    """跨工作簿复制 sheet（值与基础样式）。"""
    target = target_wb.create_sheet(title=title)
    target.sheet_format = copy(source.sheet_format)
    target.sheet_properties = copy(source.sheet_properties)
    target.merged_cells = copy(source.merged_cells)
    target.page_margins = copy(source.page_margins)
    target.freeze_panes = source.freeze_panes

    for row in source.iter_rows():
        for cell in row:
            new_cell = target[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    for col, dim in source.column_dimensions.items():
        target.column_dimensions[col].width = dim.width
    for row_idx, dim in source.row_dimensions.items():
        target.row_dimensions[row_idx].height = dim.height
    return target


def find_source_workbooks(indir: Path) -> list[Path]:
    """合并版源表：优先 indir，缺失时回退 _merged_backup。"""
    found: dict[str, Path] = {}
    for folder in (indir, indir / "_merged_backup"):
        if not folder.is_dir():
            continue
        for path in sorted(folder.glob("*.xlsx")):
            if path.name.startswith("~$") or path.name.startswith("z_"):
                continue
            found.setdefault(path.name, path)
    return [found[name] for name in sorted(found)]


def collect_groups(indir: Path) -> dict[str, dict]:
    """扫描 indir，返回 {group_key: {workbook_path, z_sheets[], extra_sheets[]}}。"""
    groups: dict[str, dict] = {}
    workbooks: dict[Path, list[str]] = {}

    for path in find_source_workbooks(indir):
        wb = load_workbook(path, read_only=True, data_only=False)
        workbooks[path] = wb.sheetnames
        wb.close()

    for path, sheet_names in workbooks.items():
        z_sheets = {name for name in sheet_names if name.startswith("z_")}
        extra = [name for name in sheet_names if not name.startswith("z_")]
        for sheet_name in sorted(z_sheets):
            key = group_key(sheet_name, z_sheets)
            if key is None:
                continue
            entry = groups.setdefault(
                key,
                {"sources": {}, "z_sheets": [], "extra_sheets": set()},
            )
            entry["z_sheets"].append((path, sheet_name))
            entry["sources"][path] = sheet_names
            for extra_name in extra:
                entry["extra_sheets"].add((path, extra_name))
    return groups


def split_indir(indir: Path, backup_dir: Path | None, dry_run: bool = False) -> list[Path]:
    groups = collect_groups(indir)
    written: list[Path] = []

    loaded: dict[Path, object] = {}
    try:
        for key in sorted(groups):
            entry = groups[key]
            output = indir / f"z_{key}.xlsx"
            sheet_plan: list[tuple[Path, str]] = []
            seen_extra: set[tuple[Path, str]] = set()
            for src_path, sheet_name in sorted(entry["z_sheets"], key=lambda item: item[1]):
                sheet_plan.append((src_path, sheet_name))
            for src_path, sheet_name in sorted(entry["extra_sheets"]):
                if (src_path, sheet_name) not in seen_extra:
                    sheet_plan.append((src_path, sheet_name))
                    seen_extra.add((src_path, sheet_name))

            if dry_run:
                print(f"[dry-run] {output.name} <- {len(sheet_plan)} sheet(s)")
                for src_path, sheet_name in sheet_plan:
                    print(f"  {sheet_name} ({src_path.name})")
                written.append(output)
                continue

            dst_wb = Workbook()
            dst_wb.remove(dst_wb.active)
            for src_path, sheet_name in sheet_plan:
                if src_path not in loaded:
                    loaded[src_path] = load_workbook(src_path, data_only=False)
                src_wb = loaded[src_path]
                copy_worksheet(src_wb[sheet_name], dst_wb, sheet_name)
            dst_wb.save(output)
            dst_wb.close()
            written.append(output)
            print(output)
    finally:
        for wb in loaded.values():
            wb.close()

    # 全部写出后再备份原合并版，避免读源文件时路径已失效
    if not dry_run and backup_dir is not None:
        backup_dir.mkdir(parents=True, exist_ok=True)
        for path in sorted(indir.glob("*.xlsx")):
            if path.name.startswith("~$") or path.name.startswith("z_"):
                continue
            shutil.move(str(path), str(backup_dir / path.name))
            print(f"backup: {path.name} -> {backup_dir.name}/")

    return written


def main() -> int:
    tool_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="按 z_{xxx} 分组拆分 Excel 配置表。")
    parser.add_argument(
        "indir",
        nargs="?",
        type=Path,
        default=tool_dir / "indir",
        help="输入目录（默认 ./indir）",
    )
    parser.add_argument(
        "--backup-dir",
        type=Path,
        default=None,
        help="合并版 xlsx 备份目录（默认 indir/_merged_backup）",
    )
    parser.add_argument("--dry-run", action="store_true", help="仅打印拆分计划，不写文件")
    args = parser.parse_args()

    indir = args.indir.resolve()
    if not indir.is_dir():
        print(f"目录不存在: {indir}", file=sys.stderr)
        return 1

    backup_dir = None if args.dry_run else (args.backup_dir or indir / "_merged_backup").resolve()
    written = split_indir(indir, backup_dir, dry_run=args.dry_run)
    action = "将生成" if args.dry_run else "已生成"
    print(f"{action} {len(written)} 个分组工作簿")
    if backup_dir is not None and not args.dry_run:
        print(f"原合并文件已移至 {backup_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
